'''
    This script takes a payroll file and creates the individual payslips. 3 files are produced:
    1. faslips.txt = Fair Acres Ltd payslips
    2. fahslips.txt = Fairacres Hortec Ltd payslips
    3. fahoslips.txt = Fair Acres Home Owners Management Lts payslips

    A different script will take each of the payslip files and send them via SMS to the individuals.

    v2.0 March 2023
'''
from openpyxl import load_workbook
import sqlite3
from database import get_db, log
import os
from datetime import datetime

# name of the workbook. This name should be standardized.
# wbname = 'PaySlips_JUL2023.xlsx'
# The payroll period is in A2 in each sheet.
# payroll_period = 'July 2023'

# load the workbook, calculate formulae (don't load formula as formula, but as calculated data)
# wb = load_workbook(wbname, data_only=True)
# print(wb.sheetnames)

# the name of the worksheet - each company is in a separate worksheet. Note that the company names are
# also in cell A1 of each sheet.
sheet_names = ['FAIR ACRES LTD', 'FAIR ACRES HOME OWNERS LTD', 'FAIRACRES HORTEC LTD']
# for sheet in sheet_names:
#     print(sheet)

sheet = None
# max_row = total rows (up to the last row) with data
# max_col = total number of columns (up to last column) with data
# max_row = sheet.max_row
# max_col = sheet.max_column

# print("max row = {} and max col = {}".format(max_row, max_col))

# initialize ranges for the rows for each of the companies. The end number is the
# number AFTER the last row in the set.
fa_range = (6,19)
fahm_range = (6, 10)
fah_range = (6, 25)

def getpayfile(fileid):
    db = get_db()
    cur = db.cursor()
    sql = 'select filename from payuploads where id = ?'
    cur.execute(sql, [fileid])
    data = cur.fetchone()
    if data is None:
        return None
    filename = data['filename']
    if os.path.isfile(filename):
        return filename
    
    return None

# return number to 2 decimal places. What we are passing is normally a string.
def fixNumber(num):
    if num is None:
        return None
    
    if isinstance(num, (float, int)):
        num = str(num)
        
    num = num.strip()
    return "{:.2f}".format(float(num))

def getSlip(company,row):
    global sheet
    '''
    Excel file has the following columns: (Note A=1, B=2, C=3,...)
        A ( 1): Employee Number
        B ( 2): Employee Full Name
        C ( 3): Phone Number (format = 2547NNNNNNNN) - if missing, no payslip is generated
        D ( 4): National ID
        E ( 5): KRA PIN
        F ( 6): Job Description
        G ( 7): Gross Pay
        H ( 8): House Allowance (not used)
        I ( 9): Bonus (not used)
        J (10): Overtime
        K (11): Benefits (?)
        L (12): NSSF (National Social Security Fund)
        M (13): Taxable Income
        N (14): NHIF (National Health Insurance Fund)
        O (15): PAYE Calc 1
        P (16): PAYE Calc 2
        Q (17): PAYE Calc 3
        R (18): PAYE
        S (19): FAWA Deduction
        T (20): Advances
        U (21): Absenteeism Deduction
        V (22): FAWA Contribution
        W (23): Housing Benefit
        X (24): Other Deductions
        Y (25): Net Pay
    '''
    slip = {
        'company': company,
        'employeeno': sheet.cell(row,1).value,
        'fullname': sheet.cell(row,2).value,
        'phone': sheet.cell(row,3).value,
        'nationalid': sheet.cell(row,4).value,
        'krapin': sheet.cell(row,5).value,
        'jobdescription': sheet.cell(row,6).value,
        'grosspay': fixNumber(sheet.cell(row,7).value),
        'houseallowance': fixNumber(sheet.cell(row,8).value),
        'otherpay': fixNumber(sheet.cell(row,9).value),
        'overtime': fixNumber(sheet.cell(row,10).value),
        'benefits': fixNumber(sheet.cell(row,11).value),
        'nssf': fixNumber(sheet.cell(row,12).value),
        'taxableincome': fixNumber(sheet.cell(row,13).value),
        'nhif': fixNumber(sheet.cell(row,14).value),
        'paye1': fixNumber(sheet.cell(row,15).value),
        'paye2': fixNumber(sheet.cell(row,16).value),
        'paye3': fixNumber(sheet.cell(row,17).value),
        'paye': fixNumber(sheet.cell(row,18).value),
        'housinglevy': fixNumber(sheet.cell(row,19).value),
        'fawaloan': fixNumber(sheet.cell(row,20).value),
        'payadvance': sheet.cell(row,21).value,
        'absent': fixNumber(sheet.cell(row,22).value),
        'fawacontribution': fixNumber(sheet.cell(row,23).value),
        'otherdeductions': fixNumber(sheet.cell(row,24).value),
        'housingbenefit': fixNumber(sheet.cell(row,25).value),
        'netpay': fixNumber(sheet.cell(row,26).value)
    }
    # phone must be included, so we give it a dummy phone that we can check later.
    if slip['phone'] is None or float(slip['phone']) == 0:
        slip['phone'] = '254000000000'
    # nationalid must not be blank also.
    return slip

def makepayrollinsert(slip, payrollheaderid):
    sql = '''
        insert into payroll (payrollid, paymonth, payyear, company, employeeno, fullname, 
        phone, nationalid, krapin, jobdescription, grosspay, houseallowance, otherpay, overtime, 
        benefits, nssf, taxableincome, nhif, paye1, paye2, paye3, paye, housinglevy, fawaloan, 
        payadvance, absent, fawacontribution, housingbenefit, otherdeductions, netpay)
        values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    '''
    return sql

def printSlip(slip, payroll_period):
    s = f'Pay Slip: {slip["company"]}\n'
    s += f'Payroll: {payroll_period}\n'
    s += slip['fullname'] + '\n'
    if slip['krapin'] is not None:
        s += f'KRA PIN: {slip["krapin"]}\n'
    s += 'Gross Pay: {:,.2f}\n'.format(float(slip["grosspay"]))
    if slip['overtime'] is not None and float(slip['overtime']) > 0:
        s += "Overtime: {:,.2f}\n".format(float(slip['overtime']))
    if slip['nssf'] is not None and float(slip['nssf']) > 0:
        s += "NSSF: {:,.2f}\n".format(float(slip['nssf']))
    if slip['nhif'] is not None and float(slip['nhif']) > 0:
        s += "NHIF: {:,.2f}\n".format(float(slip['nhif']))
    if slip['paye'] is not None and float(slip['paye']) > 0:
        s += "PAYE: {:,.2f}\n".format(float(slip['paye']))
    if slip['housinglevy'] is not None and float(slip['housinglevy']) > 0:
        s += "Housing Levy: {:,.2f}\n".format(float(slip['housinglevy']))
    if slip['fawaloan'] is not None and float(slip['fawaloan']) > 0:
        s += "FAWA Loan Repayment: {:,.2f}\n".format(float(slip['fawaloan']))
    if slip['payadvance'] is not None and float(slip['payadvance']) > 0:
        s += "Advances: {:,.2f}\n".format(float(slip['payadvance']))
    if slip['absent'] is not None and float(slip['absent']) > 0:
        s += "Absenteeism: {:,.2f}\n".format(float(slip['absent']))    
    if slip['otherdeductions'] is not None and float(slip['otherdeductions']) > 0:
        s += "Other Deductions: {:,.2f}\n".format(float(slip['otherdeductions']))    
    if slip['fawacontribution'] is not None and float(slip['fawacontribution']) > 0:
        s += "FAWA Contribution: {:,.2f}\n".format(float(slip['fawacontribution']))
    s += "NET PAY: {:,.2f}\n".format(float(slip['netpay']))
    s += f"PH: {slip['phone']}\n"
    s += "\nNet Pay ({:,.2f}) has been deposited to your bank account".format(float(slip['netpay']))

    return s

# print the FA staff payslips - only print the slip if there is a phone number.
# the getSlip() function returns 'NULL' in the phone number field if no phone number was
# provided.
# --------------- START OF OLD CODE BLOCK ---------------------
# f = open('faslips.txt', 'w')
# company = 'Fair Acres Ltd'
# # get the right sheet. sheet_names[0] = 'FAIR ACRES LTD'
# sheet = wb[sheet_names[0]]
# print(f'Printing company {company}')
# for row in range(fa_range[0],fa_range[1]):
#     slip = getSlip(company,row)
#     if slip['phone'] is not None:
#         f.write(printSlip(slip))
#         f.write('\n\n')
#     # print((sheet.cell(row,col).value).title(), end='')
#     # for key, value in slip.items():
#     #     if value is not None:
#     #         print(f'{key} -> {value}', end='')
#     # print("\n")
# f.close()

# # print("\n\n")
# # print the Fair Acres Management staff payslips
# f = open('fahoslips.txt','w')
# company = 'Fair Acres Home Owners Management Ltd'
# # sheet_names[1] = 'FAIR ACRES HOME OWNERS LTD'
# sheet = wb[sheet_names[1]]
# print(f'Printing company {company}')
# for row in range(fahm_range[0], fahm_range[1]):
#     # print((sheet.cell(row,col).value).title(), end='')
#     slip = getSlip(company,row)
#     f.write(printSlip(slip))
#     f.write('\n\n')
#     # print((sheet.cell(row,col).value).title(), end='')
#     # for key, value in slip.items():
#     #     if value is not None:
#     #         print(f'{key} -> {value}', end='')
#     # print("\n")
# f.close()

# # print("\n\n")
# # print the Fairacres Hortec staff payslips
# f = open('fahslips.txt', 'w')
# company = 'Fairacres Hortec Ltd'
# # sheet_names[2] = 'FAIRACRES HORTEC LTD'
# sheet = wb[sheet_names[2]]
# print(f'Printing company {company}')
# for row in range(fah_range[0], fah_range[1]):
#     # print((sheet.cell(row,col).value).title(), end='')
#     slip = getSlip(company,row)
#     f.write(printSlip(slip))
#     f.write('\n\n')
#     # print((sheet.cell(row,col).value).title(), end='')
#     # for key, value in slip.items():
#     #     if value is not None:
#     #         print(f'{key} -> {value}', end='')
#     # print("\n")
# f.close()

# print("\n\nDone! looks like everything completed ok.")
# -------------- END OF OLD CODE BLOCK ----------

# ----------- sqlite processing via fasms application ---------------
def checkheadings(headings):
    check_headings = [
        'EMPLOYEE NUMBER', 'EMPLOYEE', 'PHONE NUMBER', 'NATIONAL ID', 'KRA PIN', 'JOB DESCRIPTION', 
        'GROSS PAY', 'HOUSE ALLOWANCE', 'OTHER PAY', 'OVERTIME', 'BENEFITS', 'NSSF', 'TAXABLE INCOME', 'NHIF',
        '0 - 24000 (10%)', '24001 - 32333 (25%)', 'Over 32333 (30%)',
        'PAYE', 'HOUSING LEVY', 'FAWA LOAN DEDUCTION', 'ADVANCES', 'ABSENTEEISM DEDUCTION', 'FAWA CONTRIBUTION',
        'HOUSING BENEFIT', 'OTHER DEDUCTIONS', 'NET PAY'
    ]
    # headings is a dictionary structure... 
    result = True
    # just check the values...
    headings_values = headings.values()
    for v in headings_values:
        if v not in check_headings:
            log(f'checkheadings: cannot find heading {v}')
            result = False

    return result


# check that the format of the file passed is a payroll file - is valid.
def checkpayfile(fileid):
    global sheet_names, sheet

    filename = getpayfile(fileid)

    if filename is None:
        log('checkpayfile: unable to get filename')
        return False

    log(f'checkpayfile: checking file = {filename}')
    wbname = filename
    wb = None
    try:
        wb = load_workbook(wbname, data_only=True)
    except:
        log('checkpayfile: unable to load workbook using openpyxl')
        return False
    # assume first row = 6.
    # read 50 rows.
    # for each row, first column = employeeno - if employeeno not in file, record an error and continue.
    # 1 = check for sheets.
    result = True
    check_sheets = wb.sheetnames
    for asheet in sheet_names:
        if asheet not in check_sheets:
            log(f'checkpayfile: cannot find sheet {asheet}')
            result = False

    if result == False:
        return result
    
    # check each column to make sure that the data is OK.
    # this is done by checking row = 5 (titles)
    for asheet in sheet_names:
        log(f'checkpayfile: checking sheet {asheet}')
        sheet = wb[asheet]
        company = sheet.cell(1,1).value
        payperiod = sheet.cell(2,1).value
        row = 5
        slip_heading = {
            'employeeno': sheet.cell(row,1).value,
            'fullname': sheet.cell(row,2).value,
            'phone': sheet.cell(row,3).value,
            'nationalid': sheet.cell(row,4).value,
            'krapin': sheet.cell(row,5).value,
            'jobdescription': sheet.cell(row,6).value,
            'grosspay': sheet.cell(row,7).value,
            'houseallowance': sheet.cell(row,8).value,
            'otherpay': sheet.cell(row,9).value,
            'overtime': sheet.cell(row,10).value,
            'benefits': sheet.cell(row,11).value,
            'nssf': sheet.cell(row,12).value,
            'taxableincome': sheet.cell(row,13).value,
            'nhif': sheet.cell(row,14).value,
            'paye1': sheet.cell(row,15).value,
            'paye2': sheet.cell(row,16).value,
            'paye3': sheet.cell(row,17).value,
            'paye': sheet.cell(row,18).value,
            'housinglevy': sheet.cell(row,19).value,
            'fawaloan': sheet.cell(row,20).value,
            'advances': sheet.cell(row,21).value,
            'absent': sheet.cell(row,22).value,
            'fawacontribution': sheet.cell(row,23).value,
            'housingbenefit': sheet.cell(row,24).value,
            'otherdeductions': sheet.cell(row,25).value,
            'netpay': sheet.cell(row,26).value
        }
        if checkheadings(slip_heading) == False:
            log('checkpayfile: call to checkheadings returned False')
            result = False
        
    return result

def getpaymonthno(m):
    if m == 'JANUARY':
        return 1
    elif m == 'FEBRUARY':
        return 2
    elif m == 'MARCH':
        return 3
    elif m == 'APRIL':
        return 4
    elif m == 'MAY':
        return 5
    elif m == 'JUNE':
        return 6
    elif m == 'JULY':
        return 7
    elif m == 'AUGUST':
        return 8
    elif m == 'SEPTEMBER':
        return 9
    elif m == 'OCTOBER':
        return 10
    elif m == 'NOVEMBER':
        return 11
    elif m == 'DECEMBER':
        return 12
    else:
        return None

def fullmonth(m):
    if m == 'JAN':
        return 'January'
    elif m == 'FEB':
        return 'February'
    elif m == 'MAR':
        return 'March'
    elif m == 'APR':
        return 'April'
    elif m == 'JUN':
        return 'JUne'
    elif m == 'JUL':
        return 'July'
    elif m == 'AUG':
        return 'August'
    elif m == 'SEP':
        return 'September'
    elif m == 'OCT':
        return 'October'
    elif m == 'NOV':
        return 'November'
    elif m == 'DEC':
        return 'December'
    else:
        return 'Unknown' 

def getpayrolldate(d):
    data = d.split()
    if len(data) != 2:
        return None
    return data

def getprocessdate(d):
    if len(d) != 9:
        return None
    
    day = None
    try:
        day = int(d[:2])
    except:
        return None

    month = fullmonth(d[2:5])

    year = None
    try:
        year = int(d[5:])
    except:
        return None

    return list([day, month, year])

# import payroll data. 
# returns the number of rows imported.
def importpaydata(fileid):
    global sheet_names, sheet

    filename = getpayfile(fileid)
    if filename is None:
        log('importpaydata: unable to get filename')
        return 0
    if not checkpayfile(fileid):
        log(f'importpaydata: this does not look like a payroll file = {filename}')
        return 0
    
    # file is OK, now check that this pay file does not already exist in the system.
    # create the payroll header.
    # name of company is in (col,row) = (1,1)
    # payroll date is in (col,row) = (1,2)
    wb = None
    try:
        wb = load_workbook(filename, data_only=True)
    except:
        log('importpaydata: unable to load workbook using openpyxl')
        return 0
    
    log('importpaydata: workbook loaded')
    db = get_db()
    cur = db.cursor()
    rows_imported = 0
    for asheet in sheet_names:
        sheet = wb[asheet]
        log(f'reading sheet = {asheet}')
        company = sheet.cell(1,2).value
        sql = 'select id from company where company = ?'
        cur.execute(sql, [company])
        data = cur.fetchone()
        if data is None:
            continue
        companyid = int(data['id'])
        payrolldate = getpayrolldate(sheet.cell(2,2).value)
        processdate = getprocessdate(sheet.cell(3,2).value)
        if processdate is None or payrolldate is None:
            # skip this sheet if the data is incorrect.
            log('importpaydata: problem with processdate or payrolldate on Excel sheet')
            continue
        # process the file.
        # header is:
        # id, paydate, payyear, paymonth. unique(paymonth, payyear)
        paymonth = payrolldate[0]
        paymonthno = getpaymonthno(paymonth.upper())
        payyear = int(payrolldate[1])
        paydate = datetime.strptime(' '.join([str(x) for x in processdate]), '%d %B %Y')
        # check if this date/year is in the payrollheader.
        sql = 'select payid from payrollheader where companyid = ? and paymonth = ? and payyear = ?'
        cur.execute(sql, [companyid, paymonthno, payyear])
        data = cur.fetchone()
        if data is not None:
            continue
        sql = 'insert into payrollheader (companyid, paydate, paymonth, payyear) values (?, ?, ?, ?)'
        cur.execute(sql, [companyid, paydate, paymonthno, payyear])
        db.commit()
        payrollheaderid = cur.lastrowid
        log(f'importpaydata: header written id = {payrollheaderid}')
        # starting from row = 6, read a maximum of 30 rows (max number of employees in Hortec)
        row = 6
        while row < 36:
            slip = getSlip(company,row)
            employeeno = slip['employeeno']
            if employeeno is not None and employeeno[0] == 'F':
                sqlvalues = list([payrollheaderid, paymonthno, payyear]) + list(slip.values())
                sql = makepayrollinsert(slip, payrollheaderid)
                cur.execute(sql, sqlvalues)
                db.commit()
                rows_imported += 1
            row += 1
        
    return rows_imported