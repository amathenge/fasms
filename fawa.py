# helper funtions.
from openpyxl import load_workbook
import sqlite3
from database import get_db, log

# some global data for this function.
wb = None
sheet = None

# check format of string is DDMMMYYYY
def isfawasheet(sheetname):
    if len(sheetname) != 9:
        return False
    day = sheetname[:2]
    month = sheetname[2:5]
    year = sheetname[5:]
    try:
        day = int(day)
        year = int(year)
    except:
        return False
    
    if month not in ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']:
        return False

    if year < 2023:
        return False

    return True

def nullify(item):
    if item is None:
        return 'NULL'

    if isinstance(item, (float, int)):
        if item == 0:
            return 'NULL'
        else:
            return item

    if isinstance(item, str):
        if len(item.strip()) == 0:
            return 'NULL'
        else:
            return item.strip() 
    
    # we really should never get here.
    return item

def get_member(row):
    global sheet
    member = None
    data = sheet.cell(row,1).value
    try:
        memberno = int(data)
    except:
        return None
    
    # check database for this member.
    db = get_db()
    cur = db.cursor()
    sql = 'select id, memberno, firstname, middlename, surname, phone from member where memberno = ?'
    cur.execute(sql, [memberno])
    data = cur.fetchone()
    if data is None:
        return None
    
    member = {
        'memberno': memberno,
        'membername': sheet.cell(row, 2).value,
        'totaldeposit': sheet.cell(row, 3).value,
        'monthlydeposit': sheet.cell(row, 4).value,
        'totalloan_principal': nullify(sheet.cell(row, 5).value),
        'totalloanpaid': nullify(sheet.cell(row, 6).value),
        'outstandingloan': nullify(sheet.cell(row, 7).value),
        'loanrepayment': nullify(sheet.cell(row, 8).value),
        'guaranteed': nullify(sheet.cell(row, 9).value),
        'loanroom_noguarantee': nullify(sheet.cell(row, 10).value),
        'loanroom_guarantee': nullify(sheet.cell(row, 11).value),
        'phone': nullify(sheet.cell(row, 12).value)
    }

    return member

def month_number(month):
    if month == 'JAN': 
        return 1
    elif month == 'FEB':
        return 2
    elif month == 'MAR':
        return 3
    elif month == 'APR':
        return 4
    elif month == 'MAY':
        return 5
    elif month == 'JUN':
        return 6
    elif month == 'JUL':
        return 7
    elif month == 'AUG':
        return 8
    elif month == 'SEP':
        return 9
    elif month == 'OCT':
        return 10
    elif month == 'NOV':
        return 11
    elif month == 'DEC':
        return 12
    else:
        return 0
    
# the fawaheader table contains a header for the statement for a particular date.
# in this case, month/year combination.
def get_header(fileid, sheet_name):
    # assume that the sheetname is OK. DDMMMYYYY
    # return the id of the header record in the table fawaheader.
    day = int(sheet_name[:2])
    month = month_number(sheet_name[2:5])
    year = int(sheet_name[5:])
    db = get_db()
    cur = db.cursor()
    sql = '''
        select id, fileid, statementday, statementmonth, statementyear from fawaheader where 
        statementday = ? and statementmonth = ? and statementyear = ?
    '''
    cur.execute(sql, [day, month, year])
    data = cur.fetchone()
    if data is not None:
        return None
    
    sql = '''
        insert into fawaheader (fileid, statementday, statementmonth, statementyear)
        values (?, ?, ?, ?)
    '''
    cur.execute(sql, [fileid, day, month, year])
    db.commit()
    header = cur.lastrowid
    return header

# given a member object, build a query string, inserting NULL where appropriate.
def buildQuery(header, data):
    # statementid, memberno, msendgenericsmsembername cannot be null
    queryString = '(?, ?, ?'
    queryData = [header, data['memberno'], data['membername']]
    if data['totaldeposit'] == 'NULL':
        queryString += ', NULL'
    else:
        queryString += ', ?'
        queryData += ["{:.2f}".format(float(data['totaldeposit']))]
    if data['monthlydeposit'] == 'NULL':
        queryString += ', NULL'
    else:
        queryString += ', ?'
        queryData += ["{:.2f}".format(float(data['monthlydeposit']))]
    if data['totalloan_principal'] == 'NULL':
        queryString += ', NULL'
    else:
        queryString += ', ?'
        queryData += ["{:.2f}".format(float(data['totalloan_principal']))]
    if data['totalloanpaid'] == 'NULL':
        queryString += ', NULL'
    else:
        queryString += ', ?'
        queryData += ["{:.2f}".format(float(data['totalloanpaid']))]
    if data['outstandingloan'] == 'NULL':
        queryString += ', NULL'
    else:
        queryString += ', ?'
        queryData += ["{:.2f}".format(float(data['outstandingloan']))]
    if data['loanrepayment'] == 'NULL':
        queryString += ', NULL'
    else:
        queryString += ', ?'
        queryData += ["{:.2f}".format(float(data['loanrepayment']))]
    if data['guaranteed'] == 'NULL':
        queryString += ', NULL'
    else:
        queryString += ', ?'
        queryData += ["{:.2f}".format(float(data['guaranteed']))]
    if data['loanroom_noguarantee'] == 'NULL':
        queryString += ', NULL'
    else:
        queryString += ', ?'
        queryData += ["{:.2f}".format(float(data['loanroom_noguarantee']))]
    if data['loanroom_guarantee'] == 'NULL':
        queryString += ', NULL'
    else:
        queryString += ', ?'
        queryData += ["{:.2f}".format(float(data['loanroom_guarantee']))]
    # phone must be there
    queryString += ', ?)'
    queryData += [data['phone']]

    queryItems = {
        'queryString': queryString,
        'queryData': queryData
    }

    return queryItems

# given the excel file - read the sheets and insert data from the sheets into the database.
# sheets are of the format, DDMMMYYYY where DD = Day, MMM = Month, YYYY = Year.
# if the month/year combination already exists in the database, do not import it.
# returns:
#   None = failed to import
#   Integer = number of rows imported.
def importdata(fileid, filename):
    # we are modifying the global sheet and workbook variables.
    global sheet, wb

    try:
        wb = load_workbook(filename, data_only=True)
        log(f'workbook opened = {filename}')
    except:
        log(f'failed to open workbook = {filename}')
        return None

    rows_imported = 0
    sheet_names = wb.sheetnames
    log(f'Sheets are: {sheet_names}')
    # for each sheet, if the sheet has the format MMYYYY, then this is a FAWA Statement sheet.
    # read the sheet.
    db = get_db()
    cur = db.cursor()
    for each_sheet in sheet_names:
        if not isfawasheet(each_sheet):
            continue
        
        sheet = wb[each_sheet]
        header = get_header(fileid, each_sheet)
        # start at row=1 and go to about 100.
        # if col=1 has an integer, then this is a member
        cur_row = 1
        while cur_row < 100:
            member = get_member(cur_row)
            if member is not None:
                # create the insert statement for the member and add them to the database.
                # table = fawastatement
                sql = '''
                    insert into fawastatement (statementid, memberno, membername, totaldeposit,
                    monthlydeposit, totalloan_principal, totalloanpaid, outstandingloan, 
                    loanrepayment, guaranteed, loanroom_noguarantee, loanroom_guarantee, phone)
                    values
                '''
                # using the member object, create the variables to be included in the query
                queryVars = buildQuery(header, member)
                sql += ' ' + queryVars['queryString']
                queryArgs = queryVars['queryData']
                cur.execute(sql, queryArgs)
                db.commit()
                rows_imported += 1
            cur_row += 1
        log(f'imported rows = {rows_imported}')
    return rows_imported

# return a date string given day, month and year
# e.g., buildDateString(d, m, y) --> "2 January 2023"
def buildDateString(d, m, y):
    s = str(d)
    if m == 1:
        s += ' January '
    elif m == 2:
        s += ' February '
    elif m == 3:
        s += ' March '
    elif m == 4:
        s += ' April '
    elif m == 5:
        s += ' May '
    elif m == 6:
        s += ' June '
    elif m == 7:
        s += ' July '
    elif m == 8:
        s += ' August '
    elif m == 9:
        s += ' September '
    elif m == 10:
        s += ' October '
    elif m == 11:
        s += 'November'
    elif m == 12:
        s += 'December'
    else:
        s += ' UNKNOWN '

    s += str(y)

    return s